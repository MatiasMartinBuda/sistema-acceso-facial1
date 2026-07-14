"""
Módulo de Reportes (Sección 6 y 10 del PDF), exportado como .xlsx.

Genera un archivo Excel con 3 hojas:
- Resumen: totales de la última semana, por camino de decisión.
- Personas: listado completo de personas registradas.
- Logs: últimos eventos de acceso (permitidos/denegados).

El archivo se guarda en la carpeta reportes/ en la raíz del proyecto.
"""
import os
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import config
import database

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _escribir_encabezado(ws, columnas):
    ws.append(columnas)
    for celda in ws[1]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
        celda.alignment = Alignment(horizontal="center")


def _autoajustar_columnas(ws):
    for col in ws.columns:
        largo = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(largo + 2, 10), 45)


def generar_reporte_xlsx():
    os.makedirs(config.REPORTES_DIR, exist_ok=True)

    wb = Workbook()

    # --- Hoja Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    r = database.reporte_semanal()
    ws_resumen.append(["Reporte del sistema de acceso facial"])
    ws_resumen["A1"].font = Font(bold=True, size=14)
    ws_resumen.append([f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_resumen.append([])
    ws_resumen.append(["Métrica (últimos 7 días)", "Valor"])
    for celda in ws_resumen[4]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
    ws_resumen.append(["Total de eventos", r["total"]])
    ws_resumen.append(["Accesos permitidos", r["permitidos"]])
    ws_resumen.append(["Accesos denegados", r["denegados"]])
    ws_resumen.append([])
    ws_resumen.append(["Camino de decisión", "Cantidad"])
    for celda in ws_resumen[9]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
    nombres_camino = {"A": "A - Reconocimiento directo", "B": "B - Residente con PIN", "C": "C - Visita"}
    for camino, cant in r["por_camino"].items():
        ws_resumen.append([nombres_camino.get(camino, camino), cant])
    _autoajustar_columnas(ws_resumen)

    # --- Hoja Personas ---
    ws_personas = wb.create_sheet("Personas")
    _escribir_encabezado(ws_personas, [
        "ID", "DNI", "Nombre", "Apellido", "Categoría", "Depto",
        "Tipo de acceso", "Email", "PIN", "Lista negra", "Fecha de alta"
    ])
    for p in database.listar_personas():
        ws_personas.append([
            p["label_lbph"], p["dni"], p["nombre"], p["apellido"], p["categoria"],
            p["depto"], p["tipo_acceso"], p["email"] or "", p["pin"] or "",
            "Sí" if p["lista_negra"] else "No", p["fecha_alta"],
        ])
    _autoajustar_columnas(ws_personas)

    # --- Hoja Logs ---
    ws_logs = wb.create_sheet("Logs")
    _escribir_encabezado(ws_logs, [
        "Fecha/Hora", "Camino", "Resultado", "Score", "Depto destino",
        "DNI declarado", "Persona", "Detalle"
    ])
    for log in database.listar_logs_recientes(limite=1000):
        persona = f"{log['nombre']} {log['apellido']}" if log["nombre"] else ""
        ws_logs.append([
            log["timestamp"], log["camino"], log["resultado"], log["score"],
            log["depto_destino"] or "", log["dni_declarado"] or "", persona,
            log["detalle"] or "",
        ])
    _autoajustar_columnas(ws_logs)

    nombre_archivo = f"reporte_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = os.path.join(config.REPORTES_DIR, nombre_archivo)
    wb.save(ruta)
    return ruta
